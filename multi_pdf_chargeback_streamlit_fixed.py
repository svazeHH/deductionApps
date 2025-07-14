
import streamlit as st
import pdfplumber
import pandas as pd
import re
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_text_from_pdf(pdf_path):
    all_text = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)
        return "\n".join(all_text)
    except Exception as e:
        st.error(f"Error extracting text from PDF: {e}")
        return ""

def parse_pdf_content(content):
    lines = content.split('\n')
    title = next((line for line in lines if "HARMLESS HARVEST" in line), "HARMLESS HARVEST")
    week_ending = next((line for line in lines if "Week ending" in line), "")
    main_data = []
    headers = ["Brand", "Product", "Unit", "Description", "Invoice", "Ordered", "Shipped", "Wholesale", "Discount%", "MCB%", "MCB", "Customer ID", "Customer Name", "Location"]
    main_data.append(headers)
    current_location = current_customer_id = current_customer_name = ""

    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
        if re.match(r'^[A-Za-z]+\s+[A-Z]{2}$', line) and "Customer" not in line:
            current_location = line
            continue
        if line.startswith("Customer :"):
            customer_match = re.search(r'Customer : \[(\d+)\]-(.*)', line)
            if customer_match:
                current_customer_id = customer_match.group(1)
                current_customer_name = customer_match.group(2).strip()
            continue
        if line.startswith("*HRMLSHRVS"):
            try:
                parts = line.split()
                if len(parts) < 11:
                    continue
                brand = parts[0]
                product = parts[1]
                unit = f"{parts[2]} {parts[3]}"
                desc_index = 4
                description_parts = []
                while desc_index < len(parts) and not re.match(r'^\d{6,}$', parts[desc_index]):
                    description_parts.append(parts[desc_index])
                    desc_index += 1
                description = " ".join(description_parts)
                if desc_index >= len(parts) - 5:
                    invoice_match = re.search(r'(\d{8,9})', line)
                    if invoice_match:
                        invoice = invoice_match.group(1)
                        remaining = line[line.index(invoice) + len(invoice):].strip()
                        num_parts = remaining.split()
                        if len(num_parts) >= 5:
                            ordered = int(num_parts[0])
                            shipped = int(num_parts[1])
                            wholesale = float(num_parts[2])
                            discount = num_parts[3]
                            mcb_percent = num_parts[4]
                            mcb = float(num_parts[5]) if len(num_parts) > 5 else 0.0
                        else:
                            continue
                    else:
                        continue
                else:
                    invoice = parts[desc_index]
                    desc_index += 1
                    ordered = int(parts[desc_index]) if desc_index < len(parts) else 0
                    desc_index += 1
                    shipped = int(parts[desc_index]) if desc_index < len(parts) else 0
                    desc_index += 1
                    wholesale = float(parts[desc_index]) if desc_index < len(parts) else 0.0
                    desc_index += 1
                    discount = parts[desc_index] if desc_index < len(parts) else "0%"
                    desc_index += 1
                    mcb_percent = parts[desc_index] if desc_index < len(parts) else "0%"
                    desc_index += 1
                    mcb = float(parts[desc_index]) if desc_index < len(parts) else 0.0
                main_data.append([brand, product, unit, description, invoice, ordered, shipped, wholesale, discount, mcb_percent, mcb, current_customer_id, current_customer_name, current_location])
            except:
                continue
    return {
        'title': title,
        'week_ending': week_ending,
        'main_data': main_data
    }

def create_summary_tables(main_data):
    data_rows = main_data[1:]
    columns = main_data[0]
    df = pd.DataFrame(data_rows, columns=columns)
    for col in ['Ordered', 'Shipped']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    for col in ['Wholesale', 'MCB']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(float)

    location_summary = df.groupby('Location').agg({'Shipped': 'sum', 'MCB': 'sum'}).reset_index()
    location_summary.columns = ['Location', 'Total Items', 'Total MCB']
    location_summary = [location_summary.columns.tolist()] + location_summary.values.tolist()

    customer_summary = df.groupby(['Customer ID', 'Customer Name', 'Location']).agg({'Shipped': 'sum', 'MCB': 'sum'}).reset_index()
    customer_summary.columns = ['Customer ID', 'Customer Name', 'Location', 'Total Items', 'Total MCB']
    customer_summary = [customer_summary.columns.tolist()] + customer_summary.values.tolist()

    product_summary = df.groupby(['Product', 'Description']).agg({'Shipped': 'sum', 'MCB': 'sum'}).reset_index()
    product_summary.columns = ['Product', 'Description', 'Total Items', 'Total MCB']
    product_summary = [product_summary.columns.tolist()] + product_summary.values.tolist()

    return {
        'location_summary': location_summary,
        'customer_summary': customer_summary,
        'product_summary': product_summary
    }

def save_to_excel(data):
    output = io.BytesIO()
    wb = Workbook()
    def style_sheet(ws, table, currency_cols=None):
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDEBF7", fill_type="solid")
        alignment = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for r, row in enumerate(table):
            for c, val in enumerate(row):
                cell = ws.cell(row=r+4, column=c+1, value=val)
                if r == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                cell.border = border
                if r > 0 and currency_cols and c in currency_cols:
                    cell.number_format = '$#,##0.00'
        for col in range(1, len(table[0]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    ws1 = wb.active
    ws1.title = "Sales Data"
    ws1.cell(row=1, column=1, value=data['title']).font = Font(bold=True, size=14)
    ws1.cell(row=2, column=1, value=data['week_ending']).font = Font(italic=True)
    style_sheet(ws1, data['main_data'], currency_cols=[7, 10])

    for name, key, currency_cols in [
        ("Location Summary", "location_summary", [2]),
        ("Customer Summary", "customer_summary", [4]),
        ("Product Summary", "product_summary", [3])
    ]:
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value=data['title']).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=data['week_ending']).font = Font(italic=True)
        style_sheet(ws, data[key], currency_cols)

    wb.save(output)
    output.seek(0)
    return output

# Streamlit UI
st.title("🧾 Convert Multiple Chargeback PDFs to Excel")
uploaded_files = st.file_uploader("Upload one or more PDF files", type=["pdf"], accept_multiple_files=True)

if uploaded_files:
    for uploaded_file in uploaded_files:
        st.markdown(f"#### Processing: {uploaded_file.name}")
        with st.spinner("Extracting..."):
            import tempfile
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, uploaded_file.name)
        with open(temp_path, "wb") as f:
             f.write(uploaded_file.getbuffer())
        text = extract_text_from_pdf(temp_path)
        parsed = parse_pdf_content(text)
        summaries = create_summary_tables(parsed['main_data'])
        parsed.update(summaries)
        excel_data = save_to_excel(parsed)
        st.download_button(
            label=f"📥 Download Excel for {uploaded_file.name}",
            data=excel_data,
            file_name=uploaded_file.name.replace(".pdf", ".xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
